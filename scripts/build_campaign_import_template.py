"""Build frontend/campaign-import-template.xlsx with dependent dropdowns.

CampaignType → SubCampaignType, and Type (Asset/Tactic) → LineName.
Import still sends the same text fields the API already stores.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / 'frontend' / 'campaign-import-template.xlsx'
MAX_ROW = 500

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
METRICS = ['Plan', 'Forecast', 'Commit', 'Actual']

HEADER = (
    ['Action', 'Team', 'SubTeam', 'Campaign', 'SubCampaignType', 'CampaignType',
     'CampaignId', 'Type', 'LineName', 'Region', 'Products', 'Theme', 'Objective',
     'StartDate', 'EndDate', 'ConvRate', 'Year']
    + [f'{m} {met}' for m in MONTHS for met in METRICS]
    + ['Comments']
)

CAMPAIGN_TYPE_SUBTYPES = {
    'Webinar': ['Webinar'],
    'Event': ['Virtual Event', 'In-Person Event', 'Executive Dinner', 'Workshop',
              'Customer Event', 'Partner Event', 'Third-Party Event'],
    'Email': ['Email Campaign', 'Nurture Campaign'],
    'Paid': ['Paid Search', 'Paid Social', 'Display'],
    'Organic': ['Organic Social', 'Organic Search'],
    'Content Syndication': ['Content Syndication'],
    'Website Gated': ['Demo Request'],
    'Referral': ['Referral Traffic'],
    'Direct Mail': ['Direct Mail'],
}

ASSET_LINE_NAMES = [
    '1:1 Meeting', 'Agency Retainer', 'Battle Cards', 'Blog', 'Booth', 'Case Studies',
    'Contact Request', 'Demand Gen Page', 'eBook', 'Employment Guide', 'Event Registrations',
    'Event Sponsorship', 'Goodies/Swag', 'Guide', 'Infographics', 'Inquiry', 'Interactive Asset',
    'Interactive Demos', 'Landing Page Copy', 'Marketing Emails', 'Marketplace Signups',
    'Newsletter', 'On-Demand Webinar', 'Other Miscellaneous', 'Partner Lead Submission',
    'Podcast', 'Presentation', 'Promotion fee for ISV/Channel Partner', 'Reports',
    'ROI Calculator', 'Survey', 'Translations & Localizations', 'Trial Signups', 'Video',
    'Web Content Management', 'Whitepaper', 'Travel & Accom', 'Activations', 'Others',
]

TACTIC_LINE_NAMES = [
    'AE Emails', 'Agency Retainer', 'Chat', 'Content Syndication',
    'Customer Retention & Growth', 'Direct Mails', 'Direct Traffic', 'Display',
    'External List Buys', 'In-Product', 'Influencer', 'Internal Referral',
    'Marketing Emails', 'Newsletter', 'Organic Search', 'Organic Social',
    'Outbound Prospecting', 'Paid Search', 'Paid Media', 'Paid Social',
    'Partner Referral', 'Podcast', 'Press Release', 'Print', 'Referral Traffic',
    'Reseller Registrations', 'Review Site Traffic', 'SDR Emails', 'Support Email',
    'Telemarketing', 'Web Publishing',
]


def excel_name(label: str) -> str:
    return label.replace(' ', '_')


def col_index(name: str) -> int:
    return HEADER.index(name) + 1


def col_letter(name: str) -> str:
    return get_column_letter(col_index(name))


def main():
    wb = Workbook()

    lists = wb.active
    lists.title = 'Lists'
    lists['A1'] = 'CampaignType'
    lists['B1'] = 'Type'
    lists['C1'] = 'Action'
    lists['D1'] = 'Asset'
    lists['E1'] = 'Tactic'
    for i, name in enumerate(CAMPAIGN_TYPE_SUBTYPES, start=2):
        lists.cell(i, 1, name)
    for i, name in enumerate(['Campaign', 'Asset', 'Tactic'], start=2):
        lists.cell(i, 2, name)
    for i, name in enumerate(['Create', 'Update', 'Delete'], start=2):
        lists.cell(i, 3, name)
    for i, name in enumerate(ASSET_LINE_NAMES, start=2):
        lists.cell(i, 4, name)
    for i, name in enumerate(TACTIC_LINE_NAMES, start=2):
        lists.cell(i, 5, name)

    # One column per campaign type, values = its subtypes (for INDIRECT named ranges).
    subtype_start_col = 7
    for offset, (ctype, subs) in enumerate(CAMPAIGN_TYPE_SUBTYPES.items()):
        col = subtype_start_col + offset
        lists.cell(1, col, ctype)
        for i, sub in enumerate(subs, start=2):
            lists.cell(i, col, sub)
        last = 1 + len(subs)
        letter = get_column_letter(col)
        wb.defined_names.add(DefinedName(
            name=excel_name(ctype),
            attr_text=f'Lists!${letter}$2:${letter}${last}',
        ))

    n_types = len(CAMPAIGN_TYPE_SUBTYPES)
    n_assets = len(ASSET_LINE_NAMES)
    n_tactics = len(TACTIC_LINE_NAMES)
    wb.defined_names.add(DefinedName('CampaignTypeList', attr_text=f'Lists!$A$2:$A${1+n_types}'))
    wb.defined_names.add(DefinedName('TypeList', attr_text='Lists!$B$2:$B$4'))
    wb.defined_names.add(DefinedName('ActionList', attr_text='Lists!$C$2:$C$4'))
    wb.defined_names.add(DefinedName('Asset', attr_text=f'Lists!$D$2:$D${1+n_assets}'))
    wb.defined_names.add(DefinedName('Tactic', attr_text=f'Lists!$E$2:$E${1+n_tactics}'))
    # Type=Campaign rows should not pick a LineName.
    lists['F1'] = 'Campaign'
    lists['F2'] = ''
    wb.defined_names.add(DefinedName('Campaign', attr_text='Lists!$F$2:$F$2'))

    lists.sheet_state = 'hidden'

    ws = wb.create_sheet('Import', 0)
    header_fill = PatternFill('solid', fgColor='F76918')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Border(
        left=Side(style='thin', color='D6D0C2'),
        right=Side(style='thin', color='D6D0C2'),
        top=Side(style='thin', color='D6D0C2'),
        bottom=Side(style='thin', color='D6D0C2'),
    )
    for i, name in enumerate(HEADER, start=1):
        cell = ws.cell(1, i, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADER))}1'
    ws.row_dimensions[1].height = 28

    widths = {
        'Action': 12, 'Team': 22, 'SubTeam': 14, 'Campaign': 32,
        'SubCampaignType': 20, 'CampaignType': 20, 'CampaignId': 16,
        'Type': 12, 'LineName': 28, 'Region': 12, 'Products': 14,
        'Theme': 22, 'Objective': 24, 'StartDate': 12, 'EndDate': 12,
        'ConvRate': 12, 'Year': 10, 'Comments': 28,
    }
    for name, width in widths.items():
        ws.column_dimensions[col_letter(name)].width = width
    for m in MONTHS:
        for met in METRICS:
            ws.column_dimensions[col_letter(f'{m} {met}')].width = 12

    sample = [
        {
            'Action': 'Create', 'Team': 'Growth Marketing', 'SubTeam': 'Paid',
            'Campaign': 'EOR SEA Launch Example', 'SubCampaignType': 'Paid Search',
            'CampaignType': 'Paid', 'Type': 'Campaign', 'Region': 'APAC',
            'Products': 'EOR;AOR', 'Theme': 'How the World Works',
            'Objective': 'Acquisition (New Business)', 'StartDate': '2026-08-01',
            'EndDate': '2026-09-30', 'ConvRate': '0.78',
            'Comments': 'Example row - delete me',
        },
        {
            'Action': 'Create', 'Campaign': 'EOR SEA Launch Example',
            'SubCampaignType': 'Paid Search', 'CampaignType': 'Paid',
            'Type': 'Asset', 'LineName': 'Landing Page Copy', 'Year': 2026,
            'Aug Plan': 5000, 'Aug Forecast': 5000, 'Aug Commit': 4800,
            'Sep Plan': 5200, 'Sep Forecast': 5200,
        },
        {
            'Action': 'Create', 'Team': 'Field Marketing', 'SubTeam': 'FM-APAC',
            'Campaign': 'Field Event APAC Example', 'SubCampaignType': 'Third-Party Event',
            'CampaignType': 'Event', 'Type': 'Campaign', 'Region': 'APAC',
            'StartDate': '2026-10-01', 'EndDate': '2026-10-03',
            'Comments': 'Example row - delete me',
        },
        {
            'Campaign': 'Field Event APAC Example', 'Type': 'Asset',
            'LineName': 'Event Sponsorship', 'Year': 2026,
            'Oct Plan': 15000, 'Oct Forecast': 15000, 'Oct Commit': 14500,
        },
        {
            'Campaign': 'Field Event APAC Example', 'Type': 'Asset',
            'LineName': 'Booth', 'Year': 2026,
            'Oct Plan': 8000, 'Oct Forecast': 8000, 'Oct Commit': 7600,
        },
        {
            'Campaign': 'Field Event APAC Example', 'Type': 'Asset',
            'LineName': 'Travel & Accom', 'Year': 2026,
            'Oct Plan': 4000, 'Oct Forecast': 4000,
        },
    ]
    for r, row in enumerate(sample, start=2):
        for key, val in row.items():
            ws.cell(r, col_index(key), val)

    action_col = col_letter('Action')
    type_col = col_letter('Type')
    line_col = col_letter('LineName')
    ctype_col = col_letter('CampaignType')
    subtype_col = col_letter('SubCampaignType')

    def add_list_dv(formula, cells, prompt, error):
        dv = DataValidation(
            type='list', formula1=formula, allow_blank=True,
            showDropDown=False, showErrorMessage=True, showInputMessage=True,
            errorTitle='Invalid value', error=error,
            promptTitle='Choose from list', prompt=prompt,
        )
        dv.add(cells)
        ws.add_data_validation(dv)

    add_list_dv('=ActionList', f'{action_col}2:{action_col}{MAX_ROW}',
                'Create, Update, or Delete', 'Pick Create, Update, or Delete')
    add_list_dv('=CampaignTypeList', f'{ctype_col}2:{ctype_col}{MAX_ROW}',
                'Pick a campaign type first', 'Pick a campaign type from the list')
    add_list_dv(
        f'=INDIRECT(SUBSTITUTE(${ctype_col}2," ","_"))',
        f'{subtype_col}2:{subtype_col}{MAX_ROW}',
        'Pick a campaign type in this row first, then a matching sub type',
        'That sub type does not belong to the selected campaign type',
    )
    add_list_dv('=TypeList', f'{type_col}2:{type_col}{MAX_ROW}',
                'Campaign (metadata), Asset, or Tactic (spend line)',
                'Pick Campaign, Asset, or Tactic')
    add_list_dv(
        f'=INDIRECT(${type_col}2)',
        f'{line_col}2:{line_col}{MAX_ROW}',
        'Set Type to Asset or Tactic, then pick LineName',
        'LineName must match the Type (Asset list or Tactic list)',
    )

    ws[f'{ctype_col}1'].comment = Comment(
        'Fill CampaignType first. SubCampaignType dropdown then shows only matching values.',
        'Budget app',
    )
    ws[f'{line_col}1'].comment = Comment(
        'For Type=Asset or Type=Tactic rows. Campaign rows leave LineName blank.',
        'Budget app',
    )

    notes = wb.create_sheet('How to fill')
    notes['A1'] = 'How to fill this template'
    notes['A1'].font = Font(bold=True, size=14)
    notes.merge_cells('A3:A12')
    notes['A3'] = (
        '1. Keep the Import sheet headers as they are.\n'
        '2. One Type=Campaign row per campaign (Team, Campaign, CampaignType, SubCampaignType, …).\n'
        '3. Then one Type=Asset or Type=Tactic row per spend line. Rows link by Campaign name.\n'
        '4. Pick CampaignType first — SubCampaignType then only shows related values.\n'
        '5. Pick Type = Asset or Tactic — LineName then only shows that list.\n'
        '6. Leave LineName blank on Campaign rows.\n'
        '7. Delete the example rows before importing real data.\n'
        '8. Upload this .xlsx (or Save As CSV) in Data Entry → Bulk import.\n'
        '9. Values are saved as the same text fields as the campaign form — no extra database columns.'
    )
    notes['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    notes.column_dimensions['A'].width = 100
    notes.row_dimensions[3].height = 160

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f'Wrote {OUT}')


if __name__ == '__main__':
    main()
